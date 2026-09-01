"""
generate_test_data.py
=====================
Generates a test Excel file (test_advisory.xlsx) by sampling rows from the
real production tracker (Advisory_Tracker.xlsm or dummy_advisory.xlsx).

HOW TO USE:
    Double-click run_generate_test.bat
    OR run: .venv\\Scripts\\python.exe generate_test_data.py

OPTIONS (edit below):
    NUM_ROWS     - how many Alert rows to copy into test file (default: 20)
    SAMPLE_MODE  - "latest"  -> takes rows from the latest date batch only
                   "random"  -> takes random Alert rows from the whole sheet
                   "all"     -> takes ALL rows from the latest date batch
"""

import os
import sys
import random
import configparser
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl not installed. Run: .venv\\Scripts\\pip install openpyxl")
    sys.exit(1)

# ── OPTIONS ──────────────────────────────────────────────────────────────────
NUM_ROWS    = 20          # Number of Alert rows to sample into test file
SAMPLE_MODE = "latest"    # "latest" | "random" | "all"
OUTPUT_FILE = "test_advisory.xlsx"
SHEET_NAME  = "Advisory"
# ─────────────────────────────────────────────────────────────────────────────

def load_config():
    cfg = configparser.ConfigParser(inline_comment_prefixes=("#",))
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.ini")
    cfg.read(config_path, encoding="utf-8")
    primary  = cfg.get("excel", "workbook_file",     fallback="Advisory_Tracker.xlsm")
    fallback = cfg.get("excel", "workbook_fallback", fallback="dummy_advisory.xlsx")
    col_type = cfg.getint("columns", "advisory_type", fallback=2)
    col_date = cfg.getint("columns", "date",          fallback=5)
    return (primary if os.path.exists(primary) else fallback), SHEET_NAME, col_type, col_date

def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str) and val.strip():
        for fmt in ["%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except Exception:
                pass
    return None

def main():
    source_file, sheet_name, col_type, col_date = load_config()

    print("=" * 60)
    print("  CSU Advisory Alert - Test Data Generator")
    print("=" * 60)
    print(f"  Source   : {source_file}")
    print(f"  Mode     : {SAMPLE_MODE}  |  Max rows: {NUM_ROWS}")
    print(f"  Output   : {OUTPUT_FILE}")
    print("=" * 60)

    if not os.path.exists(source_file):
        print(f"[ERROR] Source file not found: {source_file}")
        sys.exit(1)

    print(f"[*] Loading source workbook (read-only)...")
    wb_src = openpyxl.load_workbook(source_file, data_only=True, read_only=True)

    if sheet_name not in wb_src.sheetnames:
        print(f"[ERROR] Sheet '{sheet_name}' not found in {source_file}")
        print(f"        Available sheets: {wb_src.sheetnames}")
        sys.exit(1)

    sheet = wb_src[sheet_name]

    # Read all rows into memory (needed for random sampling)
    all_rows = list(sheet.iter_rows(min_row=1, values_only=True))
    header   = all_rows[0] if all_rows else []
    data_rows = all_rows[1:]

    # Filter Alert rows only
    alert_rows = []
    latest_date = None
    for row in data_rows:
        if not row or len(row) < 2:
            continue
        row_type = str(row[col_type - 1] or "").strip().lower()
        if row_type != "alert":
            continue
        row_date = parse_date(row[col_date - 1])
        alert_rows.append((row, row_date))
        if row_date and (latest_date is None or row_date > latest_date):
            latest_date = row_date

    wb_src.close()

    print(f"[+] Found {len(alert_rows)} total Alert rows in source.")
    if latest_date:
        print(f"[+] Latest Advisory Preparation Date in source: {latest_date.strftime('%d-%b-%Y')}")

    # Select rows based on SAMPLE_MODE
    if SAMPLE_MODE == "latest":
        selected = [(r, d) for r, d in alert_rows if d == latest_date]
        print(f"[*] Mode=latest: {len(selected)} rows with date {latest_date.strftime('%d-%b-%Y')}")
    elif SAMPLE_MODE == "random":
        random.shuffle(alert_rows)
        selected = alert_rows[:NUM_ROWS]
        print(f"[*] Mode=random: sampling {len(selected)} random Alert rows from full history")
    else:  # "all"
        selected = alert_rows
        print(f"[*] Mode=all: using all {len(selected)} Alert rows")

    # Apply NUM_ROWS cap (except for "all" mode)
    if SAMPLE_MODE != "all" and len(selected) > NUM_ROWS:
        selected = selected[:NUM_ROWS]
        print(f"[*] Capped to {NUM_ROWS} rows.")

    if not selected:
        print("[WARN] No Alert rows selected. Check source file has rows with type='Alert'.")
        sys.exit(0)

    # Write output xlsx
    print(f"\n[*] Writing {len(selected)} rows to {OUTPUT_FILE}...")
    wb_out  = openpyxl.Workbook()
    ws_out  = wb_out.active
    ws_out.title = sheet_name

    # Write header row
    ws_out.append(list(header))

    # Write selected Alert rows
    for (row, _) in selected:
        ws_out.append(list(row))

    wb_out.save(OUTPUT_FILE)
    wb_out.close()

    print(f"\n[OK] Test file created: {OUTPUT_FILE}")
    print(f"     Rows written : {len(selected)} Alert rows + 1 header row")
    print(f"\n  Now run: run_test.bat   to send test emails to your test inbox.")
    print("=" * 60)

if __name__ == "__main__":
    main()
