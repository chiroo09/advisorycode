import os
import sys
import time
import hashlib
import webbrowser
from datetime import datetime
from email.message import EmailMessage
import openpyxl

# Ensure UTF-8 console output on all Windows machines to prevent charmap/cp1252 encoding crashes
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# Try importing Windows COM for Classic Outlook & Excel Macro Auto-Refresh
try:
    import win32com.client
    COM_AVAILABLE = True
except ImportError:
    COM_AVAILABLE = False

# =============================================================================
# CONFIGURATION
# =============================================================================
# Primary Tracker Workbook (.xlsm / .xlsx)
# Checks for 'Advisory_Tracker.xlsm' first, falls back to 'dummy_advisory.xlsx'
EXCEL_FILE = "Advisory_Tracker.xlsm" if os.path.exists("Advisory_Tracker.xlsm") else "dummy_advisory.xlsx"
SHEET_NAME = "Advisory"
MSG_OUTPUT_DIR = "output_msg"
TRACKING_FILE = ".processed_ids.txt"

# Automation Schedule: Set to 3 for 3-hour continuous check (0 = run once and exit)
SCHEDULE_INTERVAL_HOURS = 0

# Macro & Power Query Auto-Refresh before reading (Refreshes real-time SharePoint data)
ENABLE_EXCEL_AUTO_REFRESH = True
AUTO_REFRESH_WAIT_SECONDS = 20

# Safe Testing & Dispatch Controls:
# - Set AUTO_SEND = False to OPEN the email on screen in Outlook for manual review
# - Set AUTO_SEND = True to send immediately in the background
AUTO_SEND = True

# Maximum number of alerts to process per cycle (0 = unlimited, processes ALL new alerts found)
MAX_ALERTS_PER_CYCLE = 0

# ONE-TIME INITIAL BASELINE SEEDING:
# Set to True if your Excel sheet already has hundreds of old past alerts and you ONLY
# want to start alerting for NEW alerts added from today onwards (marks all existing as processed without sending emails).
SEED_ALL_EXISTING = False

# Recipient configuration for advisories and master trigger email
EMAIL_TO = "nagireddy-gari.shinysreeja@capgemini.com"
EMAIL_CC = "tejesh988@outlook.com"
# Optional: Specify sender or shared mailbox name (leave empty to use default logged-in Outlook profile)
EMAIL_SENT_ON_BEHALF = ""

# Excel Column Indices (1-based from tracker sheet)
COL_TITLE = 1             # Title
COL_TYPE = 2              # Advisory Type
COL_ADVISORY_NO = 4       # Advisory No (e.g. TI26-090)
COL_DATE = 5              # Advisory Preparation Date
COL_SUMMARY = 7           # Summary
COL_ATTACK_VECTOR = 8     # Attack Vector
COL_IMPACT_ANALYSIS = 9   # Impact Analysis
COL_IOCS = 10             # IOCs
COL_RECOMMENDATION = 11   # Recommendation / Vendor Solution
COL_REFERENCE = 12        # Reference
COL_CVE = 13              # CVE Id
COL_IMPACTED_ELEMENTS = 14# Impacted Elements
COL_ATTACK_TYPE = 15      # Attack type / Impact Type
COL_SEVERITY = 18         # Severity

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def get_processed_ids():
    if not os.path.exists(TRACKING_FILE):
        return set()
    with open(TRACKING_FILE, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())

def mark_as_processed(row_id):
    with open(TRACKING_FILE, "a", encoding="utf-8") as f:
        f.write(f"{row_id}\n")

def format_generated_datetime(date_val=None):
    """Generates exact current timestamp string matching template e.g. Aug 30th 2026, 23:06 (CET)."""
    now = datetime.now()
    day = now.day
    if 11 <= (day % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    day_str = f"{day:02d}{suffix}"
    return now.strftime(f"%b {day_str} %Y, %H:%M (CET)")

def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return format_generated_datetime()
    return str(val).strip()

def get_row_hash(row_values):
    row_str = "|".join(safe_str(val) for val in row_values)
    return hashlib.md5(row_str.encode("utf-8")).hexdigest()

def format_bullet_list(text):
    if not text:
        return ""
    lines = [line.strip("- *• \t\r\n") for line in text.splitlines() if line.strip()]
    if not lines:
        return text
    items = "".join(f"<li style='margin-bottom: 4px;'>{line}</li>" for line in lines)
    return f"<ul style='margin: 0; padding-left: 20px;'>{items}</ul>"

def get_severity_metrics(severity_text):
    sev = severity_text.lower()
    if "critical" in sev:
        return "#C00000", "Critical", 10, 315
    elif "high" in sev:
        return "#E26B00", "High", 8, 212
    elif "medium" in sev:
        return "#FFC000", "Medium", 5, 137
    elif "low" in sev:
        return "#00B050", "Low", 2, 62
    return "#1F4E79", severity_text, 1, 62

# =============================================================================
# THREAT METERS HD IMAGE GENERATOR (COMPATIBLE WITH ALL OUTLOOK / WEBMAIL CLIENTS)
# =============================================================================
def get_pil_font(size, bold=False):
    try:
        from PIL import ImageFont
        font_names = ["arialbd.ttf", "calibrib.ttf", "segoeuib.ttf"] if bold else ["arial.ttf", "calibri.ttf", "segoeui.ttf"]
        for name in font_names:
            try:
                return ImageFont.truetype(name, size)
            except Exception:
                pass
        return ImageFont.load_default()
    except Exception:
        return None

def hex_to_rgb(hex_str):
    hex_str = hex_str.lstrip("#")
    return tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))

def draw_gradient_rect(draw_img, x0, y0, x1, y1, top_hex, bot_hex):
    r1, g1, b1 = hex_to_rgb(top_hex)
    r2, g2, b2 = hex_to_rgb(bot_hex)
    h = max(1, y1 - y0)
    for y in range(y0, y1):
        ratio = (y - y0) / float(h)
        r = int(r1 + (r2 - r1) * ratio)
        g = int(g1 + (g2 - g1) * ratio)
        b = int(b1 + (b2 - b1) * ratio)
        draw_img.line([(x0, y), (x1, y)], fill=(r, g, b))

def generate_threat_meters_image(severity_text="Critical", out_path="threat_meters.png"):
    """Generates the dual Threat Meter gauge (Vendor Meter + GSOC Meter) matching original template."""
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        return None

    sev = severity_text.lower()
    if "critical" in sev:
        score = 10
    elif "high" in sev:
        score = 8
    elif "medium" in sev:
        score = 5
    elif "low" in sev:
        score = 2
    else:
        score = 1

    scale = 2
    img_w, img_h = 780 * scale, 135 * scale
    img = Image.new("RGB", (img_w, img_h), color="#FFFFFF")
    draw = ImageDraw.Draw(img)

    f_score = get_pil_font(13 * scale, bold=True)
    f_seg = get_pil_font(11 * scale, bold=True)
    f_axis = get_pil_font(12 * scale, bold=True)
    f_title = get_pil_font(13 * scale, bold=True)

    def draw_meter(ox, is_vendor, score_val):
        x_1 = ox + 20 * scale
        x_4 = ox + 104 * scale
        x_7 = ox + 188 * scale
        x_9 = ox + 272 * scale
        x_10 = ox + 350 * scale

        lbl1 = "Not Critical" if is_vendor else "Low"
        segs = [
            (x_1, x_4, "#0080D0", "#0055A5", lbl1),
            (x_4, x_7, "#F39C38", "#D96814", "Medium"),
            (x_7, x_9, "#D62020", "#A00808", "High"),
            (x_9, x_10, "#B83088", "#85155E", "Critical"),
        ]

        y_top, y_bot = 32 * scale, 58 * scale

        for (x0, x1, col_top, col_bot, label) in segs:
            draw_gradient_rect(draw, x0, y_top, x1, y_bot, col_top, col_bot)
            draw.rectangle([x0, y_top, x1, y_bot], outline="#000000", width=1*scale)
            
            if f_seg:
                bbox = draw.textbbox((0, 0), label, font=f_seg)
                tw = bbox[2] - bbox[0]
                th = bbox[3] - bbox[1]
                tx = x0 + (x1 - x0 - tw) / 2
                ty = y_top + (y_bot - y_top - th) / 2 - 1 * scale
                draw.text((tx, ty), label, fill="#FFFFFF", font=f_seg)

        # Pointer calculation
        if score_val <= 1:
            px = x_1
        elif score_val <= 4:
            px = x_1 + (score_val - 1) / 3.0 * (x_4 - x_1)
        elif score_val <= 7:
            px = x_4 + (score_val - 4) / 3.0 * (x_7 - x_4)
        elif score_val <= 9:
            px = x_7 + (score_val - 7) / 2.0 * (x_9 - x_7)
        else:
            px = x_9 + min(1.0, (score_val - 9) / 1.0) * (x_10 - x_9)

        # Pointer score text
        score_str = str(score_val)
        if f_score:
            s_bbox = draw.textbbox((0, 0), score_str, font=f_score)
            stw = s_bbox[2] - s_bbox[0]
            draw.text((px - stw/2, 6 * scale), score_str, fill="#000000", font=f_score)

        # Pointer triangle
        tri_pts = [(px - 8 * scale, 19 * scale), (px + 8 * scale, 19 * scale), (px, y_top)]
        draw.polygon(tri_pts, fill="#D9D9D9", outline="#222222")

        # Axis line
        axis_y = 66 * scale
        draw.line([(ox + 6 * scale, axis_y), (ox + 368 * scale, axis_y)], fill="#000000", width=2*scale)
        # Arrowhead
        draw.polygon([(ox + 366 * scale, axis_y - 4 * scale), (ox + 376 * scale, axis_y), (ox + 366 * scale, axis_y + 4 * scale)], fill="#000000")

        # Ticks and numbers
        ticks = [(x_1, "1"), (x_4, "4"), (x_7, "7"), (x_9, "9"), (x_10, "10")]
        for (tx_pos, num_str) in ticks:
            draw.line([(tx_pos, axis_y - 8 * scale), (tx_pos, axis_y + 8 * scale)], fill="#000000", width=2*scale)
            if f_axis:
                n_bbox = draw.textbbox((0, 0), num_str, font=f_axis)
                ntw = n_bbox[2] - n_bbox[0]
                draw.text((tx_pos - ntw/2, axis_y + 11 * scale), num_str, fill="#000000", font=f_axis)

        # Meter title
        title = "Vendor Meter" if is_vendor else "GSOC Meter"
        if f_title:
            t_bbox = draw.textbbox((0, 0), title, font=f_title)
            ttw = t_bbox[2] - t_bbox[0]
            draw.text((ox + 185 * scale - ttw/2, 108 * scale), title, fill="#000000", font=f_title)

    draw_meter(10 * scale, is_vendor=True, score_val=score)
    draw_meter(400 * scale, is_vendor=False, score_val=score)

    final_img = img.resize((780, 135), Image.Resampling.LANCZOS)
    final_img.save(out_path, "PNG")
    return out_path

# =============================================================================
# EXCEL MACRO & POWER QUERY AUTO-REFRESH (WINDOWS COM)
# =============================================================================
def trigger_excel_macro_refresh(file_path):
    """Triggers Excel 'Refresh All' to fetch real-time SharePoint data before scanning."""
    if not COM_AVAILABLE or not ENABLE_EXCEL_AUTO_REFRESH:
        return
    
    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        return
        
    print(f"[*] Triggering Excel Macro / Power Query 'Refresh All' on {os.path.basename(file_path)}...")
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        excel.Visible = False
        
        wb = excel.Workbooks.Open(abs_path)
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        
        print(f"[*] Waiting {AUTO_REFRESH_WAIT_SECONDS}s for real-time SharePoint synchronization...")
        time.sleep(AUTO_REFRESH_WAIT_SECONDS)
        wb.Save()
        wb.Close()
        excel.Quit()
        print("[+] Excel real-time data refreshed and saved successfully.")
    except Exception as e:
        print(f"[*] Note: Excel COM refresh skipped or not available on this host ({e}).")

# =============================================================================
# HEADER BANNER EMBEDDER (TI_BG.png with fallback & CID support)
# =============================================================================
def ensure_banner_file():
    possible_names = ["TI_BG.png", "TI_BG.jpg", "TI_BG.jpeg", "header.png", "banner.png", "Capgemini_banner.png"]
    for fname in possible_names:
        if os.path.exists(fname):
            return fname
            
    fallback_file = "TI_BG.png"
    try:
        from PIL import Image, ImageDraw
        img = Image.new("RGB", (860, 110), color="#001833")
        draw = ImageDraw.Draw(img)
        draw.rectangle([0, 0, 860, 6], fill="#0070AD")
        draw.rectangle([0, 104, 860, 110], fill="#005A9C")
        draw.text((25, 20), "Cyber Security Unit", fill="#FFFFFF")
        draw.text((25, 55), "Threat Intelligence", fill="#7EC8E3")
        img.save(fallback_file, "PNG")
        return fallback_file
    except Exception:
        return None

# =============================================================================
# HTML EMAIL TEMPLATE BUILDER (EXACT MATCHING ORIGINAL CSU ADVISORY)
# =============================================================================
def build_advisory_html(data, meter_img_path=None, banner_img_path=None, use_cid=True):
    sev_color, _, _, _ = get_severity_metrics(data["severity"])
    
    # Image Source Resolution (CID for Outlook, Base64 for Standalone Browser/Webmail)
    if use_cid:
        banner_src = "cid:header_banner"
        meter_src = "cid:threat_meter"
    else:
        banner_src = ""
        if banner_img_path and os.path.exists(banner_img_path):
            try:
                import base64
                with open(banner_img_path, "rb") as f:
                    b64_b = base64.b64encode(f.read()).decode("utf-8")
                banner_src = f"data:image/png;base64,{b64_b}"
            except Exception:
                pass
                
        meter_src = ""
        if meter_img_path and os.path.exists(meter_img_path):
            try:
                import base64
                with open(meter_img_path, "rb") as f:
                    b64_m = base64.b64encode(f.read()).decode("utf-8")
                meter_src = f"data:image/png;base64,{b64_m}"
            except Exception:
                pass

    banner_img_html = f'<img src="{banner_src}" width="860" height="125" alt="Cyber Security Unit Threat Intelligence" style="display: block; width: 860px; max-width: 100%; height: auto; border: 0; margin: 0; padding: 0;" />' if banner_src else '<div style="background-color: #001833; color: #ffffff; padding: 25px; font-size: 20px; font-weight: bold;">Cyber Security Unit - Threat Intelligence</div>'
    meter_img_html = f'<img src="{meter_src}" width="780" height="135" alt="Threat Meter" style="display: block; width: 100%; max-width: 780px; height: auto; margin: 0 auto; border: 0;" />' if meter_src else ''

    impacted_html = format_bullet_list(data["impacted_elements"])
    solution_html = format_bullet_list(data["recommendation"])
    summary_html = data["summary"].replace("\n", "<br>")
    impact_html = data["impact_analysis"].replace("\n", "<br>")

    # Generate Reference Links
    ref_raw = data.get("reference", "")
    ref_links = []
    for line in ref_raw.splitlines():
        line = line.strip()
        if line.startswith("http://") or line.startswith("https://"):
            ref_links.append(f"<a href='{line}' target='_blank' style='color: #0066cc; text-decoration: underline;'>{line}</a>")
        elif line:
            ref_links.append(line)
    ref_html = "<br>".join(ref_links) if ref_links else "Refer to official vendor portal."

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
</head>
<body style="font-family: Arial, Calibri, sans-serif; font-size: 12px; color: #000000; background-color: #e9ecef; margin: 0; padding: 15px;">
    <table align="center" width="860" cellpadding="0" cellspacing="0" style="max-width: 860px; width: 100%; margin: 0 auto; background-color: #ffffff; border: 1px solid #7F7F7F; border-collapse: collapse; font-family: Arial, Calibri, sans-serif;">
        <!-- Banner Image -->
        <tr>
            <td colspan="2" style="padding: 0; margin: 0; background-color: #001833; line-height: 0; font-size: 0; border-bottom: 1px solid #7F7F7F;">
                {banner_img_html}
            </td>
        </tr>

        <!-- Title Header Row (Merged directly in table to eliminate empty box/gap) -->
        <tr>
            <td colspan="2" bgcolor="#EBF1F5" style="background-color: #EBF1F5; text-align: center; padding: 8px 12px; border-bottom: 1px solid #7F7F7F;">
                <div style="font-size: 14.5px; color: #1F4E79; font-weight: bold; font-family: Arial, sans-serif;">Information Security Advisory - Alert</div>
                <div style="font-size: 11.5px; color: #000000; font-weight: bold; font-family: Arial, sans-serif; margin-top: 3px;">Date &amp; Time Issued: {data['date']}</div>
            </td>
        </tr>

        <!-- Main Metadata Rows (Clean, unbolded values, Arial 12px) -->
        <tr>
            <td width="22%" bgcolor="#EBF1F5" style="width: 22%; background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Advisory Number</td>
            <td width="78%" bgcolor="#FFFFFF" style="width: 78%; background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{data['advisory_no']}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Title</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{data['title']}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Impacted Elements</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{impacted_html or "Refer to technical analysis."}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Summary</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{summary_html}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Severity</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">
                <span>{data['severity']}</span>
                <span style="display: inline-block; width: 28px; height: 8px; background-color: {sev_color}; margin-left: 6px; vertical-align: middle; border-radius: 1px;"></span>
            </td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Impact Type</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{data['attack_type'] or "Arbitrary code execution"}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Impact Analysis</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{impact_html}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Vendor Solution</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{solution_html or "Apply vendor patches immediately."}</td>
        </tr>

        <!-- Threat Meter Header Row -->
        <tr>
            <td colspan="2" bgcolor="#EBF1F5" style="background-color: #EBF1F5; text-align: center; padding: 6px 12px; border: 1px solid #7F7F7F;">
                <div style="font-size: 14.5px; color: #1F4E79; font-weight: bold; font-family: Arial, sans-serif;">Threat Meter</div>
            </td>
        </tr>

        <!-- Threat Meter Gauge Row (Native HD PNG Image) -->
        <tr>
            <td colspan="2" bgcolor="#FFFFFF" align="center" style="background-color: #FFFFFF; text-align: center; padding: 14px 10px; border: 1px solid #7F7F7F;">
                {meter_img_html}
            </td>
        </tr>

        <!-- GSOC Assessment Header Row -->
        <tr>
            <td colspan="2" bgcolor="#EBF1F5" style="background-color: #EBF1F5; text-align: center; padding: 6px 12px; border: 1px solid #7F7F7F;">
                <div style="font-size: 14.5px; color: #1F4E79; font-weight: bold; font-family: Arial, sans-serif;">GSOC Assessment</div>
            </td>
        </tr>

        <!-- GSOC Assessment Details -->
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">Exploitation Probability</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{data['severity']}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">GSOC Risk Assessment</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">
                Successful exploitation of these vulnerabilities could lead to {data['attack_type'] or "Arbitrary code execution"}. Hence it has been categorized as {data['severity']}. No active exploitation detected so far.
            </td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">GSOC Recommendation</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">
                <div>Apply the latest patch released by the vendor.</div>
                <div style="margin-top: 6px;">{solution_html}</div>
                <div style="background-color: #FFFF00; border: 1px solid #cccc00; color: #000000; font-size: 11px; padding: 3px 6px; margin-top: 8px; display: inline-block;">/* Test changes on non-production systems before applying on production systems */</div>
            </td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">References</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{ref_html}</td>
        </tr>
        <tr>
            <td bgcolor="#EBF1F5" style="background-color: #EBF1F5; color: #1F4E79; font-weight: bold; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">References CVE's</td>
            <td bgcolor="#FFFFFF" style="background-color: #FFFFFF; color: #000000; font-size: 12px; padding: 6px 10px; border: 1px solid #7F7F7F; vertical-align: top;">{data['cve'] or "N/A"}</td>
        </tr>

        <!-- Capgemini CSU Footer -->
        <tr>
            <td colspan="2" style="background-color: #002855; color: #ffffff; text-align: center; padding: 8px 12px; font-size: 10.5px; line-height: 1.4; border-top: 1px solid #7F7F7F; font-family: Arial, sans-serif;">
                <div>The information contained in this message is proprietary and confidential. It is for Capgemini and its customers only.</div>
                <div>Copyright &copy; 2026. All rights reserved by Capgemini.</div>
                <div style="font-style: italic; margin-top: 2px;">Collaborative Business Experience&trade;</div>
            </td>
        </tr>
    </table>
</body>
</html>"""
    return html

# =============================================================================
# OUTLOOK .MSG & TRIGGER EMAIL DISPATCHER
# =============================================================================
def create_individual_advisory(outlook, data, output_path):
    import re
    subject = f"CSU Threat Intelligence Notification Advisory {data['advisory_no']} - {data['title']} ({data['severity'].upper()})"
    banner_file = ensure_banner_file()
    
    # 1. Generate threat meter PNG for this specific alert severity
    meter_filename = f"meter_{data['advisory_no']}.png"
    meter_file_path = os.path.join(MSG_OUTPUT_DIR, meter_filename)
    generate_threat_meters_image(data["severity"], meter_file_path)

    # 2. Extract CVE(s) for Vulnerability.txt attachment
    cve_content = data.get("cve", "").strip()
    if not cve_content:
        found_cves = re.findall(r"CVE-\d{4}-\d+", data.get("impact_analysis", "") + " " + data.get("attack_vector", "") + " " + data.get("title", ""))
        if found_cves:
            cve_content = "\n".join(dict.fromkeys(found_cves))
    
    vuln_txt_path = None
    if cve_content:
        vuln_txt_filename = f"Vulnerability_{data['advisory_no']}.txt"
        vuln_txt_path = os.path.join(MSG_OUTPUT_DIR, vuln_txt_filename)
        with open(vuln_txt_path, "w", encoding="utf-8") as vf:
            vf.write(cve_content + "\n")

    # Always generate a clean HTML version for universal browser & webmail viewing
    html_path = output_path.replace(".msg", ".html")
    html_general = build_advisory_html(data, meter_img_path=meter_file_path, banner_img_path=banner_file, use_cid=False)
    with open(html_path, "w", encoding="utf-8") as hf:
        hf.write(html_general)
    
    # 1. If Classic Outlook COM is available, create native Outlook .msg item in Read/Received Mode
    if outlook is not None:
        try:
            html_content = build_advisory_html(data, meter_img_path=meter_file_path, banner_img_path=banner_file, use_cid=True)
            mail_item = outlook.CreateItem(0)
            mail_item.Subject = subject
            mail_item.To = EMAIL_TO
            if EMAIL_CC:
                mail_item.CC = EMAIL_CC
            if EMAIL_SENT_ON_BEHALF:
                mail_item.SentOnBehalfOfName = EMAIL_SENT_ON_BEHALF
            
            # Embed banner image as inline CID attachment for Outlook Desktop (Hidden from attachment bar)
            if banner_file and os.path.exists(banner_file):
                abs_banner = os.path.abspath(banner_file)
                att_b = mail_item.Attachments.Add(abs_banner, 1, 0, "")
                att_b.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "header_banner")
                try:
                    att_b.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x7FFE000B", True)
                    att_b.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x37140003", 4)
                    att_b.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x370E001F", "image/png")
                except Exception:
                    pass
            
            # Embed threat meter image as inline CID attachment for Outlook Desktop (Hidden from attachment bar)
            if meter_file_path and os.path.exists(meter_file_path):
                abs_meter = os.path.abspath(meter_file_path)
                att_m = mail_item.Attachments.Add(abs_meter, 1, 0, "")
                att_m.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "threat_meter")
                try:
                    att_m.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x7FFE000B", True)
                    att_m.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x37140003", 4)
                    att_m.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x370E001F", "image/png")
                except Exception:
                    pass

            # Attach Vulnerability.txt as actual visible attachment
            if vuln_txt_path and os.path.exists(vuln_txt_path):
                abs_vuln = os.path.abspath(vuln_txt_path)
                att_v = mail_item.Attachments.Add(abs_vuln, 1, 1, "Vulnerability.txt")
                try:
                    att_v.DisplayName = "Vulnerability.txt"
                except Exception:
                    pass

            mail_item.HTMLBody = html_content
            
            # Set PR_MESSAGE_FLAGS (0x0E070003) = 1 (MSGFLAG_READ)
            # This turns the unsent draft into a normal 'Received' email with Reply / Reply All buttons
            try:
                mail_item.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x0E070003", 1)
            except Exception:
                pass
                
            abs_path = os.path.abspath(output_path)
            mail_item.SaveAs(abs_path, 3) # 3 = olMSG
            mail_item.Close(1)
            return abs_path
        except Exception as e:
            print(f"    [!] Classic Outlook COM notice ({e}), falling back to universal email format.")
    
    # 2. Universal RFC MIME email package (.msg & .eml for New Outlook, Web, and any system)
    msg_path = output_path if output_path.endswith(".msg") else output_path + ".msg"
    adv_email = EmailMessage()
    adv_email["Subject"] = subject
    adv_email["From"] = EMAIL_SENT_ON_BEHALF or "CSU Threat Intelligence <csu-alerts@capgemini.com>"
    adv_email["To"] = EMAIL_TO
    if EMAIL_CC:
        adv_email["Cc"] = EMAIL_CC
    adv_email["Date"] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
    adv_email.set_content(f"CSU Threat Intelligence Advisory {data['advisory_no']}: {data['title']}")
    adv_email.add_alternative(html_general, subtype="html")
    
    # Attach Vulnerability.txt if CVE is present
    if cve_content:
        adv_email.add_attachment(
            (cve_content + "\n").encode("utf-8"),
            maintype="text",
            subtype="plain",
            filename="Vulnerability.txt"
        )
    
    abs_path = os.path.abspath(msg_path)
    with open(abs_path, "wb") as f:
        f.write(adv_email.as_bytes())
        
    # Also save .eml version for default Windows Mail / New Outlook
    eml_path = output_path.replace(".msg", ".eml")
    with open(eml_path, "wb") as f:
        f.write(adv_email.as_bytes())
        
    return abs_path

def send_master_trigger_email(outlook, generated_files):
    if not generated_files:
        return

    date_str = datetime.now().strftime("%d %b %Y %H:%M")
    count = len(generated_files)
    
    file_list_html = "".join(f"<li style='margin-bottom: 5px;'><b>{os.path.basename(f)}</b></li>" for f in generated_files)
    
    subject = f"CSU Threat Intelligence Notification Advisory - {count} New Alert(s) [{date_str}]"
    
    trigger_html = f"""<!DOCTYPE html>
<html>
<body style="font-family: Arial, Calibri, sans-serif; font-size: 13px; color: #333333; line-height: 1.6;">
    <p>Hi Team,</p>
    <p>Please find attached <b>{count} new Threat Intelligence Security Advisories</b> generated by the CSU Threat Intelligence Automation Engine on <b>{date_str}</b>.</p>
    <ul style="margin: 12px 0; padding-left: 25px;">
        {file_list_html}
    </ul>
    <br>
    <p>Best regards,<br>
    <strong style="color: #1F4E79;">Cyber Security Unit (CSU)</strong><br>
    Threat Intelligence Team</p>
</body>
</html>
"""
    
    # Build the full MIME email message
    eml_msg = EmailMessage()
    eml_msg["To"] = EMAIL_TO
    if EMAIL_CC:
        eml_msg["Cc"] = EMAIL_CC
    eml_msg["Subject"] = subject
    eml_msg.set_content("Please find attached the new Threat Intelligence Security Advisories.")
    eml_msg.add_alternative(trigger_html, subtype="html")

    # Attach all generated .msg advisory files
    for f in generated_files:
        with open(f, "rb") as fp:
            file_data = fp.read()
            eml_msg.add_attachment(
                file_data, 
                maintype="application", 
                subtype="vnd.ms-outlook", 
                filename=os.path.basename(f)
            )

    # 1. OPTIONAL: Direct SMTP Sending (If smtp_config.py is configured with credentials)
    try:
        import smtp_config
        smtp_user = getattr(smtp_config, "SMTP_USERNAME", "").strip()
        smtp_pass = getattr(smtp_config, "SMTP_PASSWORD", "").strip()
        if smtp_user and smtp_pass:
            import smtplib
            eml_msg["From"] = smtp_user
            smtp_server = getattr(smtp_config, "SMTP_SERVER", "smtp-mail.outlook.com")
            smtp_port = getattr(smtp_config, "SMTP_PORT", 587)
            
            print(f"[*] Transmitting live email via SMTP ({smtp_server}:{smtp_port}) to {EMAIL_TO}...")
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(eml_msg)
            server.quit()
            print(f"\n[+] SUCCESS: Live email transmitted via SMTP to {EMAIL_TO} with {count} .msg attachment(s)!")
            return
    except ImportError:
        pass
    except Exception as e:
        print(f"[!] SMTP sending note: {e}")

    # 2. Dispatch via Classic Outlook COM if available (Production host)
    if outlook is not None:
        try:
            trigger_mail = outlook.CreateItem(0)
            trigger_mail.To = EMAIL_TO
            if EMAIL_CC:
                trigger_mail.CC = EMAIL_CC
            if EMAIL_SENT_ON_BEHALF:
                trigger_mail.SentOnBehalfOfName = EMAIL_SENT_ON_BEHALF
            trigger_mail.Subject = subject
            trigger_mail.HTMLBody = trigger_html
            for f in generated_files:
                trigger_mail.Attachments.Add(os.path.abspath(f))
            if AUTO_SEND:
                trigger_mail.Send()
                print(f"\n[+] Master trigger email SENT automatically via Outlook to {EMAIL_TO} with {count} .msg attachment(s)!")
            else:
                trigger_mail.Display()
                print(f"\n[+] Master trigger email opened in Classic Outlook with {count} .msg attachment(s)!")
            return
        except Exception as e:
            print(f"[!] Note: Classic Outlook COM dispatch: {e}")

    # 3. Universal Trigger Package (Local testing with New Outlook / Desktop Mail)
    eml_path = os.path.join(MSG_OUTPUT_DIR, "Master_Trigger_Email.eml")
    with open(eml_path, "wb") as fp:
        fp.write(eml_msg.as_bytes())

    summary_path = os.path.join(MSG_OUTPUT_DIR, "Master_Trigger_Email_Summary.html")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(trigger_html)

    print(f"\n[+] Master Trigger email generated with .msg attachments: {eml_path}")
    try:
        os.startfile(os.path.abspath(eml_path))
        print(f"[+] Opened trigger email in Outlook: {eml_path}")
    except Exception:
        webbrowser.open(os.path.abspath(summary_path))

# =============================================================================
# CYCLE EXECUTION FLOW (HANDLES 25K+ ROWS EFFICIENTLY WITH PROGRESS BAR)
# =============================================================================
def draw_progress_bar(current, total, bar_length=20):
    fraction = current / total if total > 0 else 1
    filled = int(bar_length * fraction)
    bar = "=" * filled + "." * (bar_length - filled)
    percent = int(fraction * 100)
    return f"[{bar}] {percent:>3}%"

def run_cycle():
    start_time = time.time()
    now_str = datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    print("\n" + "=" * 65)
    print(f"  CSU Threat Intelligence Alert Automation Cycle  [{now_str}]")
    print("=" * 65)
    
    ensure_dir(MSG_OUTPUT_DIR)
    processed_ids = get_processed_ids()
    
    # 1. Trigger live SharePoint macro refresh
    trigger_excel_macro_refresh(EXCEL_FILE)
    
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERROR] Excel file '{EXCEL_FILE}' not found.")
        return 0

    print(f"[*] Scanning workbook '{EXCEL_FILE}' (fast read-only mode for 25k+ rows)...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    except Exception as e:
        print(f"[ERROR] Failed to load Excel workbook: {e}")
        return 0
    
    if SHEET_NAME not in wb.sheetnames:
        print(f"[ERROR] Sheet '{SHEET_NAME}' not found in workbook.")
        wb.close()
        return 0
        
    sheet = wb[SHEET_NAME]
    
    outlook = None
    if COM_AVAILABLE:
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
        except Exception:
            outlook = None

    unprocessed_alerts = []
    total_scanned = 0
    total_alerts_found = 0
    
    # Step 1: Fast scan to identify all new Alert records
    for row in sheet.iter_rows(min_row=2, values_only=True):
        total_scanned += 1
        if not row or len(row) < 2:
            continue
            
        advisory_type = safe_str(row[COL_TYPE - 1]).lower()
        
        # TARGET ONLY ALERTS
        if advisory_type != "alert":
            continue
            
        total_alerts_found += 1
        row_hash = get_row_hash(row)
        if row_hash in processed_ids:
            continue # Already processed in previous cycle
            
        title = safe_str(row[COL_TITLE - 1])
        if not title:
            continue
            
        advisory_no = safe_str(row[COL_ADVISORY_NO - 1]) or "TI-Alert"
        
        data = {
            "title": title,
            "advisory_no": advisory_no,
            "date": format_generated_datetime(row[COL_DATE - 1]),
            "severity": safe_str(row[COL_SEVERITY - 1]) or "Medium",
            "cve": safe_str(row[COL_CVE - 1]),
            "summary": safe_str(row[COL_SUMMARY - 1]),
            "impact_analysis": safe_str(row[COL_IMPACT_ANALYSIS - 1]),
            "recommendation": safe_str(row[COL_RECOMMENDATION - 1]),
            "reference": safe_str(row[COL_REFERENCE - 1]),
            "impacted_elements": safe_str(row[COL_IMPACTED_ELEMENTS - 1]),
            "attack_type": safe_str(row[COL_ATTACK_TYPE - 1]),
            "row_hash": row_hash
        }
        unprocessed_alerts.append(data)

    wb.close()
    elapsed = time.time() - start_time
    print(f"[+] Scanned {total_scanned:,} total rows in {elapsed:.2f}s.")
    print(f"[*] Total 'Alert' rows in tracker: {total_alerts_found} | New unprocessed alerts: {len(unprocessed_alerts)}")

    if not unprocessed_alerts:
        print("\n[*] All alerts are up-to-date. No new emails to send.")
        return 0

    # Step 1.5: Handle One-Time Baseline Seeding (Skip old history)
    if SEED_ALL_EXISTING:
        print(f"\n[!] SEED MODE ACTIVE: Marking all {len(unprocessed_alerts)} existing alerts as processed...")
        for data in unprocessed_alerts:
            mark_as_processed(data["row_hash"])
        print(f"[+] All {len(unprocessed_alerts)} historical alerts recorded in {TRACKING_FILE}.")
        print("[*] Baseline seeding complete! Now set SEED_ALL_EXISTING = False to begin live alerting.")
        return 0

    # Step 1.6: Apply Batch Limit (Prevents attaching 3k files to one email)
    if MAX_ALERTS_PER_CYCLE > 0 and len(unprocessed_alerts) > MAX_ALERTS_PER_CYCLE:
        print(f"[*] Batch Limit Active: Processing first {MAX_ALERTS_PER_CYCLE} of {len(unprocessed_alerts)} new alerts.")
        print(f"    (Remaining {len(unprocessed_alerts) - MAX_ALERTS_PER_CYCLE} alerts will automatically process in subsequent cycles).")
        unprocessed_alerts = unprocessed_alerts[:MAX_ALERTS_PER_CYCLE]

    # Step 2: Process each new alert with a visual progress bar
    print(f"\n[*] Processing {len(unprocessed_alerts)} alert advisory package(s):")
    print("-" * 65)
    
    generated_files = []
    total_to_process = len(unprocessed_alerts)
    
    for idx, data in enumerate(unprocessed_alerts, start=1):
        progress = draw_progress_bar(idx, total_to_process)
        clean_title = "".join(c for c in data["title"] if c.isalnum() or c in " _-")[:35].strip()
        msg_filename = f"{data['advisory_no']}_Alert_{clean_title}.msg"
        msg_path = os.path.join(MSG_OUTPUT_DIR, msg_filename)
        
        print(f"  {progress} [{idx}/{total_to_process}] Generating: {data['advisory_no']} - {clean_title}...")
        
        try:
            saved_file = create_individual_advisory(outlook, data, msg_path)
            generated_files.append(saved_file)
            mark_as_processed(data["row_hash"])
            processed_ids.add(data["row_hash"])
        except Exception as e:
            print(f"    [!] Error creating {data['advisory_no']}: {e}")

    # Step 3: Trigger master notification email with all packages attached
    if generated_files:
        print("-" * 65)
        print(f"[+] Successfully generated {len(generated_files)} advisory package(s).")
        print(f"[*] Transmitting master trigger notification email to {EMAIL_TO}...")
        send_master_trigger_email(outlook, generated_files)
        first_file = os.path.abspath(generated_files[0])
        print(f"[*] Opening preview: {first_file}")
        webbrowser.open(first_file)

    return len(generated_files)

def main():
    if SCHEDULE_INTERVAL_HOURS > 0:
        print(f"[*] Starting continuous automation engine (recurring every {SCHEDULE_INTERVAL_HOURS} hour(s))...")
        print("[*] Press Ctrl+C at any time to stop.\n")
        try:
            while True:
                run_cycle()
                next_check = datetime.fromtimestamp(time.time() + SCHEDULE_INTERVAL_HOURS * 3600).strftime("%H:%M:%S")
                print(f"\n" + "=" * 65)
                print(f"[*] Cycle finished. Next automated check at: {next_check} (sleeping {SCHEDULE_INTERVAL_HOURS}h)")
                print(f"[*] Keep this window open or minimized to run continuously.")
                print("=" * 65 + "\n")
                time.sleep(SCHEDULE_INTERVAL_HOURS * 3600)
        except KeyboardInterrupt:
            print("\n[*] Automation scheduler stopped by user.")
    else:
        run_cycle()
        print("\n" + "=" * 65)
        print("  Automation cycle completed.")
        print("=" * 65)

if __name__ == "__main__":
    main()
